import openpyxl #Libreria para abrir archivos tipo xlsx
 
book = openpyxl.load_workbook('Base de datos.xlsx') #se carga el archivo de la base de datos

sheet = book.active # Se ubica en la hoja activa

a1 = sheet['A1']
a2 = sheet['B1']
a3 = sheet['C1']
a4 = sheet['D1']
a5 = sheet['E1']

domain = str(input("Ingrese el Dominio que quiere buscar: ")) #Se recibe el input del dominio

# Se crean las variables categ y conf, que son Categoría y confiabilidad del dominio respectivamente

categ = "" 
conf = ""

# Se realiza la iteración sobre la columna de los dominios para verificar si el dominio se encuentra o no en nuestra Base de Datos

for i in range (2,107):
    if sheet.cell(i,5).value == domain:
        categ = sheet.cell(i,1).value
        if sheet.cell(i,2).value == 1:
            conf = "Malicioso"
            print("El dominio "+ domain + " pertenece a la categoria "+ categ+ " con nivel de confiabilidad "+ conf)
        elif sheet.cell(i,3).value == 1:
            conf = "Poco Confiable"
            print("El dominio "+ domain + " pertenece a la categoria "+ categ+ " con nivel de confiabilidad "+ conf)
        elif sheet.cell(i,4).value == 1:
            conf = "Confiable"
            print("El dominio "+ domain + " pertenece a la categoria "+ categ+ " con nivel de confiabilidad "+ conf)

# Finalmente se realiza el Print con las caracteristicas del dominio mencionadas anteriormente